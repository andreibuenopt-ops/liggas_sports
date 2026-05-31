import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from utils.logica import calcular_classificacao, get_bracket, artilheiros_campeonato
from utils.database import listar_jogos, listar_times, get_campeonato

# ── Helpers ───────────────────────────────────────────────────
def _hex_to_openpyxl(hex_color):
    return hex_color.lstrip("#").upper()

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_fill(hex_="1A3A5C"):
    return PatternFill("solid", start_color=hex_, end_color=hex_)

def _alt_fill():
    return PatternFill("solid", start_color="F0F4F8", end_color="F0F4F8")

# ── EXCEL ─────────────────────────────────────────────────────
def exportar_excel(campeonato_id):
    camp = get_campeonato(campeonato_id)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_classificacao(wb, campeonato_id, camp)
    _sheet_jogos(wb, campeonato_id)
    _sheet_artilheiros(wb, campeonato_id)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _sheet_classificacao(wb, campeonato_id, camp):
    ws = wb.create_sheet("Classificação")
    classif = calcular_classificacao(campeonato_id)

    ws["A1"] = f"🏆 {camp['nome']} — Classificação"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = _header_fill()
    ws.merge_cells("A1:J1")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Pos", "Time", "PJ", "V", "E", "D", "GP", "GC", "SG", "PTS"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = _header_fill("2C5282")
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    for i, row in enumerate(classif, 1):
        r = i + 2
        fill = _alt_fill() if i % 2 == 0 else None
        vals = [i, row["time"], row["PJ"], row["V"], row["E"],
                row["D"], row["GP"], row["GC"], row["SG"], row["PTS"]]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = _border()
            if fill:
                cell.fill = fill
        # Destaca PTS em negrito
        ws.cell(row=r, column=10).font = Font(bold=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    for col in range(3, 11):
        ws.column_dimensions[get_column_letter(col)].width = 8

def _sheet_jogos(wb, campeonato_id):
    ws = wb.create_sheet("Jogos")
    jogos = listar_jogos(campeonato_id)
    times_map = {t["id"]: t["nome"] for t in listar_times(campeonato_id)}

    ws["A1"] = "📅 Jogos"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = _header_fill()
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Rodada", "Fase", "Time 1", "Placar 1", "Placar 2", "Time 2", "Local", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill("2C5282")
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    for i, j in enumerate(jogos, 1):
        r = i + 2
        fill = _alt_fill() if i % 2 == 0 else None
        vals = [
            j["rodada"], j["fase"],
            times_map.get(j["time1_id"], "BYE"),
            j["placar1"] if j["placar1"] is not None else "-",
            j["placar2"] if j["placar2"] is not None else "-",
            times_map.get(j["time2_id"], "BYE"),
            j["local"] or "", j["status"]
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = _border()
            if fill:
                cell.fill = fill

    widths = [8, 18, 22, 10, 10, 22, 18, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def _sheet_artilheiros(wb, campeonato_id):
    ws = wb.create_sheet("Artilheiros")
    artilheiros = artilheiros_campeonato(campeonato_id)

    ws["A1"] = "⚽ Artilheiros"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = _header_fill()
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Jogador", "Time", "Gols"], 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill("2C5282")
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    for i, a in enumerate(artilheiros, 1):
        r = i + 2
        fill = _alt_fill() if i % 2 == 0 else None
        for col, val in enumerate([a["jogador"], a["time"], a["total_gols"]], 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = _border()
            if fill:
                cell.fill = fill

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10

# ── PDF ───────────────────────────────────────────────────────
def exportar_pdf(campeonato_id):
    camp = get_campeonato(campeonato_id)
    classif = calcular_classificacao(campeonato_id)
    jogos = listar_jogos(campeonato_id)
    times_map = {t["id"]: t["nome"] for t in listar_times(campeonato_id)}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                  textColor=colors.HexColor("#1A3A5C"), fontSize=18)
    h2_style = ParagraphStyle("h2", parent=styles["Heading2"],
                               textColor=colors.HexColor("#2C5282"), fontSize=13)

    elements = []
    elements.append(Paragraph(f"🏆 {camp['nome']}", title_style))
    elements.append(Paragraph(f"Modalidade: {camp['modalidade']} | Formato: {camp['formato']}", styles["Normal"]))
    elements.append(Spacer(1, 0.5*cm))

    # Classificação
    elements.append(Paragraph("Classificação", h2_style))
    data_c = [["Pos", "Time", "PJ", "V", "E", "D", "GP", "GC", "SG", "PTS"]]
    for i, row in enumerate(classif, 1):
        data_c.append([i, row["time"], row["PJ"], row["V"], row["E"],
                        row["D"], row["GP"], row["GC"], row["SG"], row["PTS"]])

    t_c = Table(data_c, colWidths=[1.2*cm, 6*cm] + [1.5*cm]*8)
    t_c.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3A5C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
    ]))
    elements.append(t_c)
    elements.append(Spacer(1, 0.5*cm))

    # Jogos
    elements.append(Paragraph("Jogos", h2_style))
    data_j = [["Rod.", "Fase", "Time 1", "P1", "P2", "Time 2", "Status"]]
    for j in jogos:
        data_j.append([
            j["rodada"], j["fase"],
            times_map.get(j["time1_id"], "BYE"),
            j["placar1"] if j["placar1"] is not None else "-",
            j["placar2"] if j["placar2"] is not None else "-",
            times_map.get(j["time2_id"], "BYE"),
            j["status"]
        ])

    t_j = Table(data_j, colWidths=[1.2*cm, 4.5*cm, 5.5*cm, 1.5*cm, 1.5*cm, 5.5*cm, 3*cm])
    t_j.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3A5C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elements.append(t_j)

    doc.build(elements)
    buf.seek(0)
    return buf
