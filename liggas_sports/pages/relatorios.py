import streamlit as st
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from core import engine, style

def render():
    cfg = engine.get_config()
    if not cfg:
        st.info("Configure um campeonato primeiro."); return

    style.h1(st, "📄 Relatórios", "Exporte em Excel e PDF")
    style.div(st)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("""
        <div style="background:#0d1a2d;border:1px solid #1e3a5f;border-radius:8px;
             padding:22px;text-align:center;margin-bottom:10px">
            <div style="font-size:2.2rem">📊</div>
            <div style="font-family:'Barlow Condensed',sans-serif;font-size:1rem;
                 font-weight:800;letter-spacing:2px;color:#e0e6f0;margin:6px 0 3px">EXCEL</div>
            <div style="font-family:Rajdhani,sans-serif;font-size:.82rem;color:#405060">
                Classificação · Jogos · Jogadores · Artilheiros
            </div>
        </div>""", unsafe_allow_html=True)
        try:
            buf = _excel(cfg)
            st.download_button("⬇️ BAIXAR EXCEL", data=buf,
                file_name=f"{cfg['nome'].replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary")
        except Exception as e:
            st.error(f"Erro: {e}")

    with col2:
        st.markdown("""
        <div style="background:#0d1a2d;border:1px solid #1e3a5f;border-radius:8px;
             padding:22px;text-align:center;margin-bottom:10px">
            <div style="font-size:2.2rem">📄</div>
            <div style="font-family:'Barlow Condensed',sans-serif;font-size:1rem;
                 font-weight:800;letter-spacing:2px;color:#e0e6f0;margin:6px 0 3px">PDF</div>
            <div style="font-family:Rajdhani,sans-serif;font-size:.82rem;color:#405060">
                Relatório completo para impressão
            </div>
        </div>""", unsafe_allow_html=True)
        try:
            buf = _pdf(cfg)
            st.download_button("⬇️ BAIXAR PDF", data=buf,
                file_name=f"{cfg['nome'].replace(' ','_')}.pdf",
                mime="application/pdf",
                use_container_width=True, type="primary")
        except Exception as e:
            st.error(f"Erro: {e}")

def _hf(hex_): return PatternFill("solid", start_color=hex_.lstrip("#"), end_color=hex_.lstrip("#"))
def _bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _cell(ws, row, col, val, bold=False, fill=None, align="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, size=9, color="FFFFFF" if fill else "000000")
    c.alignment = Alignment(horizontal=align)
    c.border = _bdr()
    if fill: c.fill = fill
    return c

def _excel(cfg):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet_classif(wb, cfg)
    _sheet_jogos(wb)
    _sheet_jogadores(wb)
    _sheet_artilheiros(wb)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

def _sheet_classif(wb, cfg):
    ws = wb.create_sheet("Classificação")
    hdr_fill = _hf("1A3A5C")
    alt_fill = _hf("F0F4F8")
    ws["A1"] = f"🏆 {cfg['nome']} — Classificação"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = hdr_fill
    ws.merge_cells("A1:K1")
    ws["A1"].alignment = Alignment(horizontal="center")

    fmt = cfg.get("formato","")
    if "Grupo" in fmt or "Chaves" in fmt:
        grupos = engine.get_grupos_classificacao()
        row = 2
        for g, tabela in grupos.items():
            ws.cell(row=row, column=1, value=f"GRUPO {g}").font = Font(bold=True, color="FFFFFF", size=10)
            ws.cell(row=row, column=1).fill = _hf("2C5282")
            ws.merge_cells(f"A{row}:K{row}")
            row += 1
            row = _write_classif_rows(ws, tabela, row, alt_fill)
            row += 1
    else:
        tabela = engine.calcular_classificacao()
        hdrs = ["Pos","Time","J","V","E","D","GP","GC","SG","Pts","%"]
        for ci, h in enumerate(hdrs, 1):
            _cell(ws, 2, ci, h, bold=True, fill=_hf("2C5282"))
        _write_classif_rows(ws, tabela, 3, alt_fill)

    for ci, w in enumerate([6,24,6,6,6,6,6,6,7,7,8], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

def _write_classif_rows(ws, tabela, start_row, alt_fill):
    hdrs = ["Pos","Time","J","V","E","D","GP","GC","SG","Pts","%"]
    if ws.cell(row=start_row-1, column=1).value not in ["Pos", *[f"GRUPO {g}" for g in "ABCDEFGH"]]:
        for ci, h in enumerate(hdrs, 1):
            _cell(ws, start_row-1, ci, h, bold=True, fill=_hf("2C5282"))
    for i, r in enumerate(tabela, 1):
        ri = start_row + i - 1
        vals = [i, r["time"], r["J"], r["V"], r["E"], r["D"],
                r["GP"], r["GC"], r["SG"], r["P"], f"{r['pct']}%"]
        fill = alt_fill if i % 2 == 0 else None
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = Alignment(horizontal="center" if ci != 2 else "left")
            c.border = _bdr()
            if fill: c.fill = fill
        ws.cell(row=ri, column=10).font = Font(bold=True)
    return start_row + len(tabela)

def _sheet_jogos(wb):
    ws = wb.create_sheet("Jogos")
    jogos = engine.get_jogos()
    ws["A1"] = "📅 Jogos"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = _hf("1A3A5C")
    ws.merge_cells("A1:H1")
    hdrs = ["Fase","Rodada","Casa","P1","P2","Fora","Data","Local"]
    for ci, h in enumerate(hdrs, 1):
        _cell(ws, 2, ci, h, bold=True, fill=_hf("2C5282"))
    alt = _hf("F0F4F8")
    for ri, j in enumerate(jogos, 1):
        fill = alt if ri % 2 == 0 else None
        vals = [j.get("fase",""), j.get("rodada",""), j.get("casa",""),
                j.get("gols_casa","—") if j.get("gols_casa") is not None else "—",
                j.get("gols_fora","—") if j.get("gols_fora") is not None else "—",
                j.get("fora",""), j.get("data",""), j.get("local","")]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri+2, column=ci, value=v)
            c.alignment = Alignment(horizontal="center")
            c.border = _bdr()
            if fill: c.fill = fill
    for ci, w in enumerate([20,8,22,6,6,22,12,18], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

def _sheet_jogadores(wb):
    ws = wb.create_sheet("Jogadores")
    jogs = engine.get_jogadores()
    ws["A1"] = "👤 Jogadores"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = _hf("1A3A5C")
    ws.merge_cells("A1:G1")
    hdrs = ["Nome","Equipe","Posição","Número","Gols","Amarelos","Vermelhos"]
    for ci, h in enumerate(hdrs, 1):
        _cell(ws, 2, ci, h, bold=True, fill=_hf("2C5282"))
    evs = engine.get_eventos()
    for ri, j in enumerate(jogs, 1):
        gols = sum(1 for e in evs if e["jogador_id"]==j["id"] and e["tipo"]=="gol")
        am   = sum(1 for e in evs if e["jogador_id"]==j["id"] and e["tipo"]=="amarelo")
        vm   = sum(1 for e in evs if e["jogador_id"]==j["id"] and e["tipo"]=="vermelho")
        for ci, v in enumerate([j["nome"],j.get("equipe",""),j.get("posicao",""),
                                  j.get("numero",""),gols,am,vm], 1):
            c = ws.cell(row=ri+2, column=ci, value=v)
            c.alignment = Alignment(horizontal="center")
            c.border = _bdr()
    for ci, w in enumerate([24,20,16,8,6,8,8], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

def _sheet_artilheiros(wb):
    ws = wb.create_sheet("Artilheiros")
    arts = engine.artilheiros()
    ws["A1"] = "⚽ Artilheiros"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = _hf("1A3A5C")
    ws.merge_cells("A1:C1")
    for ci, h in enumerate(["Jogador","Equipe","Gols"], 1):
        _cell(ws, 2, ci, h, bold=True, fill=_hf("2C5282"))
    alt = _hf("F0F4F8")
    for ri, a in enumerate(arts, 1):
        fill = alt if ri % 2 == 0 else None
        for ci, v in enumerate([a["nome"],a["equipe"],a["gols"]], 1):
            c = ws.cell(row=ri+2, column=ci, value=v)
            c.alignment = Alignment(horizontal="center")
            c.border = _bdr()
            if fill: c.fill = fill
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 8

def _pdf(cfg):
    tabela = engine.calcular_classificacao()
    jogos = engine.get_jogos()
    arts = engine.artilheiros()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("t", parent=styles["Title"],
                              textColor=colors.HexColor("#1A3A5C"), fontSize=16)
    h2_s = ParagraphStyle("h2", parent=styles["Heading2"],
                           textColor=colors.HexColor("#2C5282"), fontSize=11)
    tstyle = TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1A3A5C")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("GRID",(0,0),(-1,-1),.5,colors.lightgrey),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F0F4F8")]),
        ("FONTSIZE",(0,0),(-1,-1),7.5),
    ])
    elements = []
    elements.append(Paragraph(f"🏆 {cfg['nome']}", title_s))
    elements.append(Paragraph(f"Formato: {cfg.get('formato','—')}", styles["Normal"]))
    elements.append(Spacer(1, .35*cm))

    elements.append(Paragraph("Classificação", h2_s))
    data_c = [["#","Time","J","V","E","D","GP","GC","SG","Pts","%"]]
    for i, r in enumerate(tabela, 1):
        data_c.append([i,r["time"],r["J"],r["V"],r["E"],r["D"],
                        r["GP"],r["GC"],f"{r['SG']:+d}",r["P"],f"{r['pct']}%"])
    t = Table(data_c, colWidths=[1*cm,6*cm]+[1.5*cm]*9)
    t.setStyle(tstyle); elements.append(t)
    elements.append(Spacer(1, .35*cm))

    elements.append(Paragraph("Jogos", h2_s))
    data_j = [["Fase","Rod.","Casa","P1","P2","Fora","Data","Local"]]
    for j in jogos[:50]:
        data_j.append([j.get("fase",""),j.get("rodada",""),j.get("casa",""),
                        j.get("gols_casa","—") if j.get("gols_casa") is not None else "—",
                        j.get("gols_fora","—") if j.get("gols_fora") is not None else "—",
                        j.get("fora",""),j.get("data",""),j.get("local","")])
    t2 = Table(data_j, colWidths=[4*cm,1.5*cm,5*cm,1.2*cm,1.2*cm,5*cm,2.5*cm,3.5*cm])
    t2.setStyle(tstyle); elements.append(t2)

    if arts:
        elements.append(Spacer(1,.3*cm))
        elements.append(Paragraph("Artilheiros", h2_s))
        data_a = [["Jogador","Equipe","Gols"]]
        for a in arts: data_a.append([a["nome"],a["equipe"],a["gols"]])
        t3 = Table(data_a, colWidths=[8*cm,8*cm,2*cm])
        t3.setStyle(tstyle); elements.append(t3)

    doc.build(elements)
    buf.seek(0); return buf
